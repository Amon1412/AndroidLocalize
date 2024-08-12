import os
from openpyxl import load_workbook, Workbook
from datetime import datetime

# 定义输入文件和输出目录
input_file = 'merged_output.xlsx'  # 请替换为你的合并后的Excel文件
output_directory = 'splitted_excels'  # 输出目录

# 如果输出目录不存在，创建它
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

# 加载合并后的工作簿
workbook = load_workbook(input_file)
sheet = workbook.active

# 初始化变量
split_workbook = None
split_sheet = None
row_counter = 1
file_name = None

for row in sheet.iter_rows(values_only=False):
    # 检查当前行是否为空行
    if all(cell.value is None for cell in row):
        # 如果当前行是空行，保存当前拆分的工作簿
        if split_workbook and file_name:
            split_workbook.save(os.path.join(output_directory, f'{file_name}.xlsx'))
        # 重置变量，开始新的拆分工作簿
        split_workbook = None
        split_sheet = None
        row_counter = 1
        file_name = None
    else:
        # 如果当前行不是空行，将其复制到当前拆分的工作簿中
        if not split_workbook:
            split_workbook = Workbook()
            split_sheet = split_workbook.active

        for col_index, cell in enumerate(row, 1):
            new_cell = split_sheet.cell(row=row_counter, column=col_index, value=cell.value)
            # 复制单元格样式
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()
        
        # 在第3行设置文件名
        if row_counter == 3:
            if split_sheet.cell(row=3, column=2).value:
                file_name = split_sheet.cell(row=3, column=2).value
            else:
                # 使用当前时间作为文件名
                file_name = datetime.now().strftime("%H%M%S")
        
        row_counter += 1

# 保存最后一个拆分的工作簿
if split_workbook and file_name:
    split_workbook.save(os.path.join(output_directory, f'{file_name}.xlsx'))

print(f'Excel文件已拆分，并保存到 {output_directory} 目录中')
