import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# 定义Excel文件所在目录和输出文件名
input_directory = 'E:\\work\\translateExcel'  # 请替换为你的Excel文件所在目录
output_file = 'merged_output.xlsx'

# 获取目录中所有Excel文件的文件名，并按名称排序
excel_files = sorted([f for f in os.listdir(input_directory) if f.endswith('.xlsx')])

# 创建一个新的工作簿
merged_workbook = Workbook()
merged_sheet = merged_workbook.active
merged_sheet.title = 'MergedSheet'

# 设置初始行号
current_row = 1

# 遍历每个Excel文件并读取数据
for excel_file in excel_files:
    file_path = os.path.join(input_directory, excel_file)
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # 复制列宽
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        merged_sheet.column_dimensions[column_letter].width = sheet.column_dimensions[column_letter].width
    
    for row in sheet.iter_rows():
        for cell in row:
            new_cell = merged_sheet.cell(row=current_row, column=cell.column, value=cell.value)
            
            # 复制单元格样式
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()

        current_row += 1
    
    # 添加一个空行作为分隔符
    current_row += 1

# 保存合并后的工作簿
merged_workbook.save(output_file)
print(f'所有Excel文件已成功合并，并保存到 {output_file}')
